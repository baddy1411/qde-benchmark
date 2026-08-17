#!/usr/bin/env python3
"""Build the defense/review results workbook from committed CSVs.

One Excel file collecting every measured run and every pre-specified statistical
comparison, with parameters attached, so a reader can check a claim without
opening the repository.

  Read me            legend, colour key, provenance
  DM comparisons     all 321 per-seed tests; the 214 pre-specified family flagged
  Measured results   947 scaling-family runs with every parameter
  Model scoreboard   360 leaderboard runs, all models, all metrics
  Summary            counts and medians, computed by formula

Colour is categorical only: quantum rows blue, classical rows grey. Anything
computed -- winners, significance, best-in-group -- is an Excel formula so the
sheet recalculates rather than carrying my arithmetic as a literal.

Run:  .venv/bin/python scripts/build_results_workbook.py
Out:  results/QDE_Results_Workbook.xlsx
"""
from __future__ import annotations

import os

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "results", "QDE_Results_Workbook.xlsx")

FONT = "Arial"
NAVY = "152245"
C_QUANTUM = "DCE9F7"     # quantum rows
C_CLASSICAL = "EDEFF2"   # classical rows
C_BEST = "FFE699"        # best value in its group
C_FLAG = "F8CBC0"        # the single reversed comparison
C_HEAD = "152245"
C_BAND = "F4F7FB"

thin = Side(style="thin", color="D5DCEA")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def head(ws, headers, widths, row=1):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=C_HEAD)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def write_rows(ws, df, start=2, fills=None, numfmt=None):
    """fills: per-row fill colour. numfmt: {column_name: format}."""
    cols = list(df.columns)
    for r, (_, rec) in enumerate(df.iterrows(), start=start):
        fill = PatternFill("solid", fgColor=fills[r - start]) if fills else None
        for i, col in enumerate(cols, start=1):
            v = rec[col]
            if isinstance(v, (np.integer,)):
                v = int(v)
            elif isinstance(v, (np.floating,)):
                v = None if pd.isna(v) else float(v)
            elif pd.isna(v):
                v = None
            c = ws.cell(row=r, column=i, value=v)
            c.font = Font(name=FONT, size=9.5, color=NAVY)
            c.border = BORDER
            if fill:
                c.fill = fill
            if numfmt and col in numfmt:
                c.number_format = numfmt[col]
            if isinstance(v, str) and len(v) < 18:
                c.alignment = Alignment(horizontal="left")
    return start + len(df) - 1


def holm(pvals):
    """Holm-Bonferroni adjusted p-values (step-down), documented in the sheet."""
    p = np.asarray(pvals, dtype=float)
    m = len(p)
    order = np.argsort(p)
    adj = np.empty(m)
    running = 0.0
    for rank, idx in enumerate(order):
        running = max(running, (m - rank) * p[idx])
        adj[idx] = min(1.0, running)
    return adj


# ---------------------------------------------------------------- load
scores = pd.read_csv(os.path.join(ROOT, "results/scaling_proof/scores.csv"),
                     float_precision="round_trip")
dm = pd.read_csv(os.path.join(ROOT, "results/scaling_proof/dm.csv"),
                 float_precision="round_trip")
board = pd.read_csv(os.path.join(ROOT, "results/model_scoreboard_all_runs.csv"),
                    float_precision="round_trip")

QUANTUM_MODELS = {"qrc"}
OPPONENT = {"qrc_vs_elmF": "ELM(F) — feature-matched control",
            "qrc_vs_ngrc": "NG-RC — strong structural reference",
            "qrc_vs_esnF": "ESN(F) — size-matched reservoir"}
FAMILY = {"qrc_vs_elmF", "qrc_vs_ngrc"}

wb = Workbook()

# ================================================================ Read me
ws = wb.active
ws.title = "Read me"
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 34
ws.column_dimensions["C"].width = 96

ws["B2"] = "QDE benchmark — measured results and statistical comparisons"
ws["B2"].font = Font(name=FONT, bold=True, size=16, color=NAVY)
ws["B3"] = ("Quantum vs classical reservoir computing on chaotic time series · "
            "M.Sc. Data Engineering · Badrish Madapuji Srinivasan")
ws["B3"].font = Font(name=FONT, size=10, color="6E7A94")

rows = [
    ("", ""),
    ("SHEETS", ""),
    ("DM comparisons",
     "All 321 per-seed Diebold-Mariano tests, with the 214-member pre-specified "
     "family flagged in the 'In 214 family' column."),
    ("   Why 214 and not 321",
     "ESN(F) carries exactly the same feature budget as the quantum model at every "
     "size (64, 96, 128, 160, 192), so it IS feature-matched -- it sits outside the "
     "statistical family, not outside the matching. The family is the confirmatory "
     "test of the research question: ELM(F) asks whether the quantum feature map "
     "beats a classical one of the same size, and NG-RC asks whether it is "
     "competitive with the strongest classical method. ESN(F) answers neither; it "
     "is simply a weak reservoir on these tasks, four to five times worse than the "
     "quantum model. The quantum model wins all 107 of those comparisons, and that "
     "result is reported rather than buried -- it is the basis of the thesis's own "
     "point that beating a size-matched baseline is not a quantum advantage, since "
     "NG-RC remains several orders of magnitude ahead on 21 features."),
    ("Measured results",
     "All 947 runs of the qubit-scaling study with every parameter that defines "
     "them: system, qubit count, feature budget F, seed, virtual nodes, window, "
     "shots and read-out set."),
    ("Model scoreboard",
     "All 360 leaderboard runs across every model family on one common split, "
     "with all seven error metrics."),
    ("Summary", "Counts and medians, computed from the other sheets by formula."),
    ("", ""),
    ("COLOUR KEY", ""),
    ("Blue rows", "Quantum model (gate-based quantum reservoir, QRC)."),
    ("Grey rows", "Classical model (ELM, ESN, NG-RC, and the wider baseline field)."),
    ("Amber cell",
     "Best value in its category — lowest error. Driven by a formula, so it "
     "follows the data if a value changes."),
    ("Coral row",
     "The single comparison of 214 that favoured the quantum model "
     "(Henon, 4 qubits, seed 0). See the note on that row."),
    ("", ""),
    ("HOW TO READ NRMSE", ""),
    ("NRMSE",
     "Root-mean-square error divided by the standard deviation of the target. "
     "Lower is better. 1.0 is the score of predicting the mean, so any value near "
     "1.0 means the model has learned essentially nothing."),
    ("DM statistic",
     "Diebold-Mariano test of equal forecast accuracy. NEGATIVE means the quantum "
     "model had the smaller loss; POSITIVE means the classical model did."),
    ("Holm-adjusted p",
     "Holm-Bonferroni step-down correction across the 214-member family, computed "
     "in Python because the step-down procedure has no clean spreadsheet form. "
     "All 214 remain significant after correction."),
    ("", ""),
    ("PROVENANCE", ""),
    ("Source artifacts",
     "results/scaling_proof/{scores,dm}.csv and results/model_scoreboard_all_runs.csv, "
     "all committed. Built by scripts/build_results_workbook.py."),
    ("Common conditions",
     "Every run: min-max scaling fitted on training data only, 60/20/20 "
     "chronological split, washout 100, horizon 1 step, ridge alpha 1e-6. "
     "Exact expectation values (no shot sampling) unless a shots column says otherwise."),
]
r = 5
for label, text in rows:
    if label and not text:
        c = ws.cell(row=r, column=2, value=label)
        c.font = Font(name=FONT, bold=True, size=10, color="DB4D3A")
    elif label:
        ws.cell(row=r, column=2, value=label).font = Font(name=FONT, bold=True, size=10, color=NAVY)
        c = ws.cell(row=r, column=3, value=text)
        c.font = Font(name=FONT, size=10, color=NAVY)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(15, 13 * (len(text) // 95 + 1))
    r += 1

for addr, colour in (("B18", C_QUANTUM), ("B19", C_CLASSICAL), ("B20", C_BEST), ("B21", C_FLAG)):
    ws[addr].fill = PatternFill("solid", fgColor=colour)

# ================================================================ DM comparisons
d = dm.copy()
d["Quantum model"] = "QRC-rich"
d["Compared against"] = d.comparison.map(OPPONENT)
d["In 214 family"] = np.where(d.comparison.isin(FAMILY), "yes", "no (extra)")
fam_mask = d.comparison.isin(FAMILY)
d["Holm-adjusted p (family only)"] = np.nan
d.loc[fam_mask, "Holm-adjusted p (family only)"] = holm(d.loc[fam_mask, "p"].values)
d = d.sort_values(["comparison", "system", "n", "seed"],
                  key=lambda s: s.map({"qrc_vs_elmF": 0, "qrc_vs_ngrc": 1, "qrc_vs_esnF": 2})
                  if s.name == "comparison" else s).reset_index(drop=True)

dv = pd.DataFrame({
    "System": d.system,
    "Qubits": d.n,
    "Quantum model": d["Quantum model"],
    "Compared against": d["Compared against"],
    "In 214 family": d["In 214 family"],
    "Seed": d.seed,
    "DM statistic": d.dm_stat,
    "p-value": d.p,
    "Holm-adjusted p (family only)": d["Holm-adjusted p (family only)"],
})

ws = wb.create_sheet("DM comparisons")
head(ws, list(dv.columns) + ["Winner", "Significant at 0.05", "Note"],
     [12, 8, 13, 34, 13, 7, 13, 13, 15, 13, 11, 46])
fills = [C_QUANTUM if v < 0 else C_CLASSICAL for v in dv["DM statistic"]]
last = write_rows(ws, dv, fills=fills,
                  numfmt={"DM statistic": "0.0000",
                          "p-value": "0.00E+00",
                          "Holm-adjusted p (family only)": "0.00E+00"})
for r in range(2, last + 1):
    ws.cell(row=r, column=10, value=f'=IF(G{r}<0,"QUANTUM","Classical")')
    ws.cell(row=r, column=11, value=f'=IF(H{r}<0.05,"yes","no")')
    note = ws.cell(row=r, column=12)
    quantum_won = ws.cell(row=r, column=7).value < 0
    in_family = ws.cell(row=r, column=5).value == "yes"
    if quantum_won and in_family:
        # exactly one row in the whole workbook
        note.value = ("THE ONE REVERSAL OF 214. The control here is a random projection "
                      "and this seed is its worst draw of twenty (0.0062 against a best "
                      "of 0.0015); the quantum model scores 0.0044 on every seed. The "
                      "Windows re-run reversed this same cell.")
        for col in range(1, 13):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=C_FLAG)
    elif quantum_won:
        note.value = ("Quantum wins. ESN(F) is equally feature-matched but is the weakest "
                      "model here, so this is expected and is reported, not hidden — it is "
                      "the basis of the thesis's point that beating a size-matched baseline "
                      "is not a quantum advantage. NG-RC beats both on 21 features.")
    for col in (10, 11, 12):
        ws.cell(row=r, column=col).font = Font(name=FONT, size=9.5, color=NAVY,
                                               bold=(col == 10))
        ws.cell(row=r, column=col).border = BORDER
ws.auto_filter.ref = f"A1:L{last}"

# ================================================================ Measured results
s = scores.copy()
s["Model type"] = np.where(s.model.isin(QUANTUM_MODELS), "Quantum", "Classical")
NAME = {"qrc": "QRC-rich (quantum)", "elmF": "ELM(F) — feature-matched",
        "esnF": "ESN(F) — size-matched", "ngrc": "NG-RC — strong reference"}
s["Model"] = s.model.map(NAME)
s = s.sort_values(["system", "n_qubits_eq", "model", "seed"]).reset_index(drop=True)

sv = pd.DataFrame({
    "System": s.system, "Qubit setting": s.n_qubits_eq, "Model": s.Model,
    "Model type": s["Model type"], "Seed": s.seed, "NRMSE": s.nrmse,
    "Features F": s.n_features, "Points": s.n_points, "Virtual nodes V": s.V,
    "Window": s.window, "Shots": s.shots, "Read-out": s.readout,
    "NG-RC degree": s.ngrc_degree,
})

ws = wb.create_sheet("Measured results")
head(ws, list(sv.columns) + ["Best in this system+size"],
     [12, 12, 25, 10, 6, 13, 10, 8, 13, 8, 8, 13, 11, 20])
fills = [C_QUANTUM if t == "Quantum" else C_CLASSICAL for t in sv["Model type"]]
last = write_rows(ws, sv, fills=fills, numfmt={"NRMSE": "0.000000"})
for r in range(2, last + 1):
    c = ws.cell(row=r, column=14)
    c.value = (f'=IF(F{r}=_xlfn.MINIFS($F$2:$F${last},$A$2:$A${last},$A{r},'
               f'$B$2:$B${last},$B{r}),"BEST","")')
    c.font = Font(name=FONT, size=9.5, bold=True, color="0A9B75")
    c.border = BORDER
    c.alignment = Alignment(horizontal="center")
ws.conditional_formatting.add(
    f"F2:F{last}",
    FormulaRule(formula=[f'$N2="BEST"'], fill=PatternFill("solid", fgColor=C_BEST), stopIfTrue=False))
ws.auto_filter.ref = f"A1:N{last}"

# ================================================================ Model scoreboard
b = board.copy()
b["Model type"] = np.where(b.group == "quantum", "Quantum", "Classical")
b = b.sort_values(["system", "nrmse"]).reset_index(drop=True)
bv = pd.DataFrame({
    "System": b.system, "Model": b.model_name, "Model type": b["Model type"],
    "Family": b.group, "Approach": b.kind, "Seed": b.seed, "NRMSE": b.nrmse,
    "RMSE": b.rmse, "MAE": b.mae, "R2": b.r2, "sMAPE": b.smape,
    "Max error": b.max_error, "Features": b.n_features, "Parameters": b.n_params,
    "Lookback": b.cfg_lookback, "Train rows": b.n_train, "Test rows": b.n_test,
})
ws = wb.create_sheet("Model scoreboard")
head(ws, list(bv.columns) + ["Best in system"],
     [12, 22, 10, 11, 40, 6, 12, 12, 12, 10, 10, 11, 9, 11, 9, 10, 9, 14])
fills = [C_QUANTUM if t == "Quantum" else C_CLASSICAL for t in bv["Model type"]]
last = write_rows(ws, bv, fills=fills,
                  numfmt={"NRMSE": "0.00E+00", "RMSE": "0.00E+00", "MAE": "0.00E+00",
                          "R2": "0.0000", "sMAPE": "0.000", "Max error": "0.00E+00"})
for r in range(2, last + 1):
    c = ws.cell(row=r, column=18)
    c.value = f'=IF(G{r}=_xlfn.MINIFS($G$2:$G${last},$A$2:$A${last},$A{r}),"BEST","")'
    c.font = Font(name=FONT, size=9.5, bold=True, color="0A9B75")
    c.border = BORDER
    c.alignment = Alignment(horizontal="center")
ws.conditional_formatting.add(
    f"G2:G{last}",
    FormulaRule(formula=['$R2="BEST"'], fill=PatternFill("solid", fgColor=C_BEST), stopIfTrue=False))
ws.auto_filter.ref = f"A1:R{last}"

# ================================================================ Summary
ws = wb.create_sheet("Summary")
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 52
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 60
ws["B2"] = "Summary — every figure below is a formula over the other sheets"
ws["B2"].font = Font(name=FONT, bold=True, size=13, color=NAVY)

n_dm = len(dv) + 1
n_sc = len(sv) + 1
items = [
    ("Pre-specified comparisons in the family",
     f"=COUNTIF('DM comparisons'!E2:E{n_dm},\"yes\")",
     "Quantum vs ELM(F) and quantum vs NG-RC, across systems, sizes and seeds."),
    ("... won by the classical model",
     f"=COUNTIFS('DM comparisons'!E2:E{n_dm},\"yes\",'DM comparisons'!G2:G{n_dm},\">0\")",
     "Positive DM statistic means the quantum model carried the larger loss."),
    ("... won by the quantum model",
     f"=COUNTIFS('DM comparisons'!E2:E{n_dm},\"yes\",'DM comparisons'!G2:G{n_dm},\"<0\")",
     "The single reversal, highlighted coral on the DM sheet."),
    ("Largest raw p-value in the family",
     f"=_xlfn.MAXIFS('DM comparisons'!H2:H{n_dm},'DM comparisons'!E2:E{n_dm},\"yes\")",
     "Every comparison in the family is significant; this is the weakest of them."),
    ("Largest Holm-adjusted p-value in the family",
     f"=MAX('DM comparisons'!I2:I{n_dm})",
     "All 214 survive correction for multiple testing."),
    ("", "", ""),
    ("Mean NRMSE, quantum (scaling study)",
     f"=AVERAGEIF('Measured results'!$D$2:$D${n_sc},\"Quantum\",'Measured results'!$F$2:$F${n_sc})",
     "Mean rather than median: a median over a filtered range needs an array "
     "formula, which does not survive being written by a script. The thesis "
     "reports medians; these are the means over the same rows."),
    ("Mean NRMSE, classical (scaling study)",
     f"=AVERAGEIF('Measured results'!$D$2:$D${n_sc},\"Classical\",'Measured results'!$F$2:$F${n_sc})",
     "Same basis, classical rows only."),
    ("Best quantum NRMSE / best classical NRMSE",
     f"=_xlfn.MINIFS('Measured results'!$F$2:$F${n_sc},'Measured results'!$D$2:$D${n_sc},\"Quantum\")"
     f"&\"  /  \"&_xlfn.MINIFS('Measured results'!$F$2:$F${n_sc},'Measured results'!$D$2:$D${n_sc},\"Classical\")",
     "The single best run of each type, anywhere in the scaling study."),
    ("Best (lowest) NRMSE anywhere in the scaling study",
     f"=MIN('Measured results'!F2:F{n_sc})", "Across all systems, sizes, models and seeds."),
    ("Measured runs recorded",
     f"=COUNTA('Measured results'!A2:A{n_sc})", "One row per model, system, size and seed."),
]
r = 4
for label, formula, note in items:
    if not label:
        r += 1
        continue
    ws.cell(row=r, column=2, value=label).font = Font(name=FONT, size=10, color=NAVY)
    c = ws.cell(row=r, column=3, value=formula)
    c.font = Font(name=FONT, size=10, bold=True, color=NAVY)
    c.fill = PatternFill("solid", fgColor=C_BAND)
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER
    if "p-value" in label:
        c.number_format = "0.00E+00"
    elif "NRMSE" in label:
        c.number_format = "0.000000"
    n = ws.cell(row=r, column=4, value=note)
    n.font = Font(name=FONT, size=9, color="6E7A94")
    n.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print(f"wrote {os.path.relpath(OUT, ROOT)}")
print(f"  DM comparisons  {len(dv)} rows ({int(fam_mask.sum())} in the pre-specified family)")
print(f"  Measured results {len(sv)} rows")
print(f"  Model scoreboard {len(bv)} rows")
